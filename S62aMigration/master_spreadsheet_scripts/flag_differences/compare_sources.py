import csv
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HORIZON_FILE = (
    "/Users/nisalihalwathura/PINS/ODW-Service/odw-synapse-workspace/S62aMigration/outputs/MASTER LEGACY cases S62A - with Horizon data.xlsx"
)

SPREADSHEET_FILE = (
    "/Users/nisalihalwathura/PINS/ODW-Service/odw-synapse-workspace/S62aMigration/outputs/S62A_All_Sheets_migrated.xlsx"
)

MASTER_TEMPLATE_FILE = (
    "/Users/nisalihalwathura/PINS/ODW-Service/odw-synapse-workspace/"
    "S62aMigration/csv_and_xlsx_files/MASTER LEGACY cases S62A .xlsx"
)

SHEET_NAME = "Template"

KEY_COL = "Case reference"

HEADER_ROW = 2

OUTPUT_FILE = (
    "/Users/nisalihalwathura/PINS/ODW-Service/"
    "odw-synapse-workspace/S62aMigration/outputs/"
    "S62A_Horizon_vs_Spreadsheet_comparison.xlsx"
)

SUMMARY_SHEET_NAME = "Contradiction Summary"

HORIZON_MAPPING_FILE = (
    "/Users/nisalihalwathura/PINS/ODW-Service/odw-synapse-workspace/"
    "S62aMigration/outputs/horizon_field_mapping.csv"
)

# highlight differences in red
DIFFERENCE_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)


def _get_extended_data_fields(mapping_file=HORIZON_MAPPING_FILE):
    with open(mapping_file, newline="", encoding="utf-8-sig") as mapping_stream:
        return {
            row["Field"].strip()
            for row in csv.DictReader(mapping_stream)
            if row.get("Field", "").strip()
            and row.get("Source field", "").strip().startswith("extended_data.")
        }


def _get_columns(worksheet):
    columns = []
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(HEADER_ROW, column).value
        if value is None or str(value).strip() == "":
            continue
        columns.append((str(value).strip(), column))
    return columns


def _normalise_case_reference(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _is_blank(value):
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return value != value


def _values_differ(left, right):
    if _is_blank(left) or _is_blank(right):
        return False
    if isinstance(left, str) and isinstance(right, str):
        return left.strip() != right.strip()
    return left != right


def _read_cases(worksheet, columns):
    cases = {}
    order = []
    source_rows = {}
    key_column = next(column for header, column in columns if header == KEY_COL)

    for row in range(HEADER_ROW + 1, worksheet.max_row + 1):
        case_reference = _normalise_case_reference(
            worksheet.cell(row, key_column).value
        )
        if case_reference is None:
            continue

        if case_reference not in cases:
            cases[case_reference] = {
                index: worksheet.cell(row, column).value
                for index, (_, column) in enumerate(columns)
            }
            order.append(case_reference)
            source_rows[case_reference] = row
            continue

        for index, (_, column) in enumerate(columns):
            existing = cases[case_reference][index]
            incoming = worksheet.cell(row, column).value
            if existing in (None, "") and incoming not in (None, ""):
                cases[case_reference][index] = incoming

    return cases, order, source_rows


def _copy_cell_format(source_cell, target_cell):
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)


def _copy_column_format(source_worksheet, source_column, target_worksheet, target_column):
    source_letter = get_column_letter(source_column)
    target_letter = get_column_letter(target_column)
    source_dimension = source_worksheet.column_dimensions[source_letter]
    target_dimension = target_worksheet.column_dimensions[target_letter]
    target_dimension.width = source_dimension.width
    target_dimension.hidden = source_dimension.hidden
    target_dimension.bestFit = source_dimension.bestFit

    for row in range(1, HEADER_ROW + 2):
        _copy_cell_format(
            source_worksheet.cell(row, source_column),
            target_worksheet.cell(row, target_column),
        )


def _write_contradiction_summary(workbook, master_headers, contradiction_counts):
    if SUMMARY_SHEET_NAME in workbook.sheetnames:
        del workbook[SUMMARY_SHEET_NAME]

    summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    summary_sheet.append(["Master template field", "Contradiction count"])

    header_fill = PatternFill(
        fill_type="solid",
        start_color="C00000",
        end_color="C00000",
    )
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = copy(header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    written_fields = set()
    for index, field_name in enumerate(master_headers):
        if field_name == KEY_COL or field_name in written_fields:
            continue
        count = sum(
            contradiction_counts.get(field_index, 0)
            for field_index, candidate_name in enumerate(master_headers)
            if candidate_name == field_name
        )
        if count == 0:
            continue
        summary_sheet.append([field_name, count])
        written_fields.add(field_name)

    summary_sheet.column_dimensions["A"].width = 48
    summary_sheet.column_dimensions["B"].width = 22
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = f"A1:B{summary_sheet.max_row}"
    summary_sheet.sheet_view.showGridLines = False


def _output_column(source_column, source_is_key=False):
    if source_is_key:
        return 1
    return 2 + ((source_column - 2) * 2)


def combine_sources(
    horizon_file=HORIZON_FILE,
    spreadsheet_file=SPREADSHEET_FILE,
    master_template_file=MASTER_TEMPLATE_FILE,
    output_file=OUTPUT_FILE,
):
    # Write one output row per case with Horizon and spreadsheet values paired
    horizon_workbook = load_workbook(horizon_file, data_only=False)
    spreadsheet_workbook = load_workbook(spreadsheet_file, data_only=False)
    master_workbook = load_workbook(master_template_file, data_only=False)

    try:
        horizon_source = horizon_workbook[SHEET_NAME]
        spreadsheet_source = spreadsheet_workbook[SHEET_NAME]
        master_source = master_workbook[SHEET_NAME]
        horizon_columns = _get_columns(horizon_source)
        spreadsheet_columns = _get_columns(spreadsheet_source)
        master_columns = _get_columns(master_source)

        horizon_headers = [header for header, _ in horizon_columns]
        spreadsheet_headers = [header for header, _ in spreadsheet_columns]
        master_headers = [header for header, _ in master_columns]
        extended_data_fields = _get_extended_data_fields()
        if KEY_COL not in horizon_headers or KEY_COL not in spreadsheet_headers:
            raise ValueError(f"Both sheets must contain {KEY_COL!r}")

        missing_horizon_headers = [
            header for header in master_headers if header not in horizon_headers
        ]
        missing_spreadsheet_headers = [
            header for header in master_headers if header not in spreadsheet_headers
        ]
        if missing_horizon_headers or missing_spreadsheet_headers:
            raise ValueError(
                "Both sources must contain every legacy master column. Missing "
                f"from Horizon: {missing_horizon_headers}; missing from spreadsheet: "
                f"{missing_spreadsheet_headers}"
            )
        spreadsheet_only_headers = [
            header for header in spreadsheet_headers if header not in master_headers
        ]

        horizon_cases, horizon_order, horizon_rows = _read_cases(
            horizon_source, horizon_columns
        )
        spreadsheet_cases, spreadsheet_order, spreadsheet_rows = _read_cases(
            spreadsheet_source, spreadsheet_columns
        )

        output_workbook = load_workbook(master_template_file, data_only=False)
        output_sheet = output_workbook[SHEET_NAME]
        try:
            source_merges = list(output_sheet.merged_cells.ranges)
            for merged_range in source_merges:
                output_sheet.unmerge_cells(str(merged_range))
            if output_sheet.max_row > HEADER_ROW:
                output_sheet.delete_rows(HEADER_ROW + 1, output_sheet.max_row - HEADER_ROW)

            key_index = master_headers.index(KEY_COL)
            output_headers = [KEY_COL]
            # Entries are (header, summary index, source index, is_horizon).
            output_sources = [(KEY_COL, key_index, key_index, False)]
            comparison_pairs = []
            for index, header in enumerate(master_headers):
                if header != KEY_COL:
                    horizon_column = len(output_headers) + 1
                    output_headers.append(f"{header} (horizon)")
                    output_sources.append((header, index, index, True))
                    if header in extended_data_fields:
                        continue
                    spreadsheet_column = len(output_headers) + 1
                    output_headers.append(f"{header} (spreadsheet)")
                    output_sources.append((header, index, index, False))
                    comparison_pairs.append((horizon_column, spreadsheet_column, index))
            for header in spreadsheet_only_headers:
                source_index = spreadsheet_headers.index(header)
                summary_index = len(master_headers)
                master_headers.append(header)
                horizon_column = len(output_headers) + 1
                output_headers.append(f"{header} (horizon)")
                output_sources.extend(
                    ((header, summary_index, None, True),)
                )
                spreadsheet_column = len(output_headers) + 1
                output_headers.append(f"{header} (spreadsheet)")
                output_sources.append((header, summary_index, source_index, False))
                comparison_pairs.append((horizon_column, spreadsheet_column, summary_index))

            output_columns_by_source_index = {}
            for output_column, (_, index, _, _) in enumerate(output_sources, start=1):
                output_columns_by_source_index.setdefault(index, []).append(output_column)

            for output_column, (_, _, source_index, is_horizon) in enumerate(
                output_sources, start=1
            ):
                if source_index is None:
                    output_sheet.cell(HEADER_ROW, output_column).value = output_headers[
                        output_column - 1
                    ]
                    continue
                source_worksheet = horizon_source if is_horizon else spreadsheet_source
                source_columns = horizon_columns if is_horizon else spreadsheet_columns
                source_column = source_columns[source_index][1]
                _copy_column_format(
                    source_worksheet,
                    source_column,
                    output_sheet,
                    output_column,
                )
                output_sheet.cell(HEADER_ROW, output_column).value = output_headers[
                    output_column - 1
                ]

            for merged_range in source_merges:
                if merged_range.min_row != 1 or merged_range.max_row != 1:
                    continue
                mapped_columns = [
                    output_column
                    for index, (_, source_column) in enumerate(master_columns)
                    if merged_range.min_col <= source_column <= merged_range.max_col
                    for output_column in output_columns_by_source_index.get(index, [])
                ]
                if not mapped_columns:
                    continue
                start_column = min(mapped_columns)
                end_column = max(mapped_columns)
                output_sheet.cell(1, start_column).value = master_source.cell(
                    1, merged_range.min_col
                ).value
                _copy_cell_format(
                    master_source.cell(1, merged_range.min_col),
                    output_sheet.cell(1, start_column),
                )
                output_sheet.merge_cells(
                    start_row=1,
                    start_column=start_column,
                    end_row=1,
                    end_column=end_column,
                )

            case_references = list(spreadsheet_order)
            case_references.extend(
                case_reference
                for case_reference in horizon_order
                if case_reference not in spreadsheet_cases
            )
            contradiction_counts = {}

            for output_row, case_reference in enumerate(
                case_references, start=HEADER_ROW + 1
            ):
                spreadsheet_row = spreadsheet_rows.get(case_reference)
                horizon_row = horizon_rows.get(case_reference)
                output_sheet.cell(output_row, 1).value = case_reference

                if spreadsheet_row is not None:
                    _copy_cell_format(
                        spreadsheet_source.cell(
                            spreadsheet_row, spreadsheet_columns[key_index][1]
                        ),
                        output_sheet.cell(output_row, 1),
                    )

                for output_column, (_, index, source_index, is_horizon) in enumerate(
                    output_sources[1:], start=2
                ):
                    cases = horizon_cases if is_horizon else spreadsheet_cases
                    source_row = horizon_row if is_horizon else spreadsheet_row
                    source_worksheet = horizon_source if is_horizon else spreadsheet_source
                    source_columns = horizon_columns if is_horizon else spreadsheet_columns
                    output_sheet.cell(output_row, output_column).value = (
                        cases.get(case_reference, {}).get(source_index)
                        if source_index is not None else None
                    )
                    if source_row is not None and source_index is not None:
                        _copy_cell_format(
                            source_worksheet.cell(source_row, source_columns[source_index][1]),
                            output_sheet.cell(output_row, output_column),
                        )

                for horizon_column, spreadsheet_column, field_index in comparison_pairs:
                    horizon_value = output_sheet.cell(
                        output_row, horizon_column
                    ).value
                    spreadsheet_value = output_sheet.cell(
                        output_row, spreadsheet_column
                    ).value
                    if _values_differ(horizon_value, spreadsheet_value):
                        contradiction_counts[field_index] = (
                            contradiction_counts.get(field_index, 0) + 1
                        )
                        output_sheet.cell(
                            output_row, horizon_column
                        ).fill = copy(DIFFERENCE_FILL)
                        output_sheet.cell(
                            output_row, spreadsheet_column
                        ).fill = copy(DIFFERENCE_FILL)

                source_row = spreadsheet_row or horizon_row
                if source_row is not None:
                    source_worksheet = (
                        spreadsheet_source if spreadsheet_row is not None else horizon_source
                    )
                    output_sheet.row_dimensions[output_row].height = source_worksheet.row_dimensions[
                        source_row
                    ].height

            output_sheet.freeze_panes = f"A{HEADER_ROW + 1}"
            output_sheet.auto_filter.ref = (
                f"A{HEADER_ROW}:{get_column_letter(len(output_headers))}"
                f"{HEADER_ROW + len(case_references)}"
            )
            _write_contradiction_summary(
                output_workbook,
                master_headers,
                contradiction_counts,
            )
            Path(output_file).parent.mkdir(parents=True, exist_ok=True)
            output_workbook.save(output_file)
        finally:
            output_workbook.close()
    finally:
        horizon_workbook.close()
        spreadsheet_workbook.close()
        master_workbook.close()


if __name__ == "__main__":
    combine_sources()
    print(f"Created {OUTPUT_FILE}")