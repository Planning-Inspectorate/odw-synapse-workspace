import csv
import os
import re
 
import openpyxl
import pandas as pd
 
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
 
def find_data_root(start_dir, marker="csv_and_xlsx_files", max_up=4):
    d = start_dir
    for _ in range(max_up + 1):
        if os.path.isdir(os.path.join(d, marker)):
            return d
        parent = os.path.dirname(d)
        if parent == d:
            break
        d = parent
    raise FileNotFoundError(
        f"Could not find a '{marker}' folder above {start_dir} "
        f"(searched {max_up + 1} levels up) - check the script's location."
    )
 
DATA_ROOT = find_data_root(BASE_DIR)
 
def find_source_file(data_root):
    ss_data_dir = os.path.join(data_root, "csv_and_xlsx_files", "SS_data")
    if not os.path.isdir(ss_data_dir):
        raise FileNotFoundError(f"SS_data folder not found: {ss_data_dir}")
    candidates = [
        f for f in os.listdir(ss_data_dir)
        if f.lower().endswith(".xlsx")
        and "62a" in f.lower()
        and "cases" in f.lower()
        and "copy" in f.lower()
    ]
    if not candidates:
        raise FileNotFoundError(
            f"Could not find a 'Section 62a Cases ... COPY.xlsx'-style file in {ss_data_dir}. "
            f"Files present: {os.listdir(ss_data_dir)}"
        )
    if len(candidates) > 1:
        print(f"WARNING: multiple candidate source files found, using the first: {candidates}")
    return os.path.join(ss_data_dir, candidates[0])
 
SOURCE_FILE   = find_source_file(DATA_ROOT)
MAPPING_XLSX  = os.path.join(DATA_ROOT, "csv_and_xlsx_files/SS_data/S62A_Column_mapping.xlsx")
MAPPING_SHEET = "Lookup"
TEMPLATE_FILE = os.path.join(DATA_ROOT, "csv_and_xlsx_files/MASTER LEGACY cases S62A .xlsx")
TEMPLATE_SHEET      = "Template"
TEMPLATE_HEADER_ROW = 2
 
OUTPUT_FILE = os.path.join(DATA_ROOT, "outputs/spreadsheet_field_mapping_validation_v2.csv")
 
os.makedirs(os.path.join(DATA_ROOT, "outputs"), exist_ok=True)
 
print("Block 1 done - config set")
 
# %%
# ============================================================
# BLOCK 2: LOAD LOOKUP MAPPING - same parsing as map_ss_onto_template_v2.py.
# Keep this in sync with that script's load_lookup_mapping() so this
# validator checks exactly what the migration actually does.
# ============================================================
 
SET_TO_BE_RE = re.compile(r'set to be\s+"([^"]+)"', re.IGNORECASE)
 
def load_lookup_mapping():
    wb = openpyxl.load_workbook(MAPPING_XLSX, data_only=True)
    ws = wb[MAPPING_SHEET]
    rows = []
    current_category = None
    for r in range(2, ws.max_row + 1):
        category = ws.cell(row=r, column=1).value
        field    = ws.cell(row=r, column=2).value
        if category:
            current_category = category
        if not field:
            continue
        rows.append({
            "category": current_category,
            "field": str(field).strip(),
            "Pre-application - DONE":  ws.cell(row=r, column=3).value,
            "Application (Major)":     ws.cell(row=r, column=4).value,
            "Application (Non Major)": ws.cell(row=r, column=5).value,
            "notes": ws.cell(row=r, column=7).value,
        })
    return rows
 
LOOKUP_ROWS = load_lookup_mapping()
 
SHEET_NAMES = ["Pre-application - DONE", "Application (Major)", "Application (Non Major)"]
 
# Fields the migration script sources from elsewhere rather than straight
# off the Lookup sheet's stated column (address/agent-email splits) - kept
# in sync with MANUAL_SOURCE_OVERRIDES in map_ss_onto_template_v2.py so
# this validator checks the real source column, not the "N/A" the Lookup
# sheet shows for these.
MANUAL_SOURCE_OVERRIDES = {
    "Site address 2":    "Address",
    "Site town or city":  "Address",
    "Site county":        "Address",
    "Site post code":     "Address",
    "Agent email":        "Agent",
}
 
# Kept in sync with SOURCE_COLUMN_ALIASES in map_ss_onto_template_v2.py.
SOURCE_COLUMN_ALIASES = {
    "Application (Non Major)": {
        "SAP5 to FSSD": "SAP5 to FSSD / Fee requested by BACS",
    },
}
 
# Kept in sync with SHEET_FIELD_SKIPS in map_ss_onto_template_v2.py.
SHEET_FIELD_SKIPS = {
    "Application (Major)": {"CIL amount"},
}
 
def _norm_col(text):
    return re.sub(r"\s+", " ", str(text).strip()).lower()
 
def resolve_source_column(sheet_name, source_col, available_by_norm, available):
    alias = SOURCE_COLUMN_ALIASES.get(sheet_name, {}).get(source_col)
    if alias:
        source_col = alias
    if source_col in available:
        return source_col
    return available_by_norm.get(_norm_col(source_col))
 
print(f"Block 2 done - {len(LOOKUP_ROWS)} template fields loaded from Lookup sheet")
 
# %%
# ============================================================
# BLOCK 3: READ REAL SOURCE COLUMNS for each sheet
# ============================================================
 
def read_columns(sheet_name):
    return pd.read_excel(SOURCE_FILE, sheet_name=sheet_name, nrows=0).columns.tolist()
 
sheet_columns = {}
for sheet_name in SHEET_NAMES:
    cols = read_columns(sheet_name)
    sheet_columns[sheet_name] = {
        "available": set(cols),
        "available_by_norm": {_norm_col(c): c for c in cols},
    }
    print(f"  {sheet_name}: {len(cols)} columns read")
 
print("Block 3 done - source columns loaded for all three sheets")
 
# %%
# ============================================================
# BLOCK 4: VALIDATE - for every field x sheet, check the stated source
# column actually resolves to a real column on that sheet
# ============================================================
 
results = []
 
for row in LOOKUP_ROWS:
    field = row["field"]
    for sheet_name in SHEET_NAMES:
        if field in SHEET_FIELD_SKIPS.get(sheet_name, set()):
            results.append({
                "Field": field, "Sheet": sheet_name, "Source field (Lookup sheet)": "(skipped)",
                "Status": "SKIPPED", "Resolved column": "",
            })
            continue
 
        if field in MANUAL_SOURCE_OVERRIDES:
            stated_source = MANUAL_SOURCE_OVERRIDES[field]
        else:
            raw = row.get(sheet_name)
            stated_source = str(raw).strip() if raw is not None else ""
 
        if not stated_source or stated_source.upper() == "N/A":
            results.append({
                "Field": field, "Sheet": sheet_name, "Source field (Lookup sheet)": "N/A",
                "Status": "NOT MAPPED (N/A)", "Resolved column": "",
            })
            continue
 
        m = SET_TO_BE_RE.search(stated_source)
        if m:
            results.append({
                "Field": field, "Sheet": sheet_name, "Source field (Lookup sheet)": stated_source,
                "Status": "CONSTANT", "Resolved column": f'constant = "{m.group(1)}"',
            })
            continue
 
        available = sheet_columns[sheet_name]["available"]
        available_by_norm = sheet_columns[sheet_name]["available_by_norm"]
        resolved = resolve_source_column(sheet_name, stated_source, available_by_norm, available)
 
        if resolved:
            results.append({
                "Field": field, "Sheet": sheet_name, "Source field (Lookup sheet)": stated_source,
                "Status": "FOUND", "Resolved column": resolved,
            })
        else:
            results.append({
                "Field": field, "Sheet": sheet_name, "Source field (Lookup sheet)": stated_source,
                "Status": "MISSING", "Resolved column": "",
            })
 
status_counts = {}
for r in results:
    status_counts[r["Status"]] = status_counts.get(r["Status"], 0) + 1
 
print("\nBlock 4 done - validation totals:")
for status, count in sorted(status_counts.items()):
    print(f"  {status}: {count}")
 
missing = [r for r in results if r["Status"] == "MISSING"]
print(f"\n{len(missing)} MISSING entries (Lookup sheet points at a column that doesn't exist on that sheet):")
for r in missing:
    print(f"  {r['Sheet']:28} {r['Field']!r:45} -> {r['Source field (Lookup sheet)']!r}")
 
# %%
# ============================================================
# BLOCK 5: CHECK FOR TEMPLATE FIELD NAME MISMATCHES
# Every Field in the Lookup sheet should exist as a real Template header
# (after DESTINATION_FIELD_RENAMES is applied) - flag any that don't.
# ============================================================
 
DESTINATION_FIELD_RENAMES = {
    "Pre-application or application": "Application phase",
}
 
wb = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)
ws = wb[TEMPLATE_SHEET]
template_headers = {
    str(ws.cell(row=TEMPLATE_HEADER_ROW, column=c).value).strip()
    for c in range(1, ws.max_column + 1)
    if ws.cell(row=TEMPLATE_HEADER_ROW, column=c).value
}
 
lookup_fields = {row["field"] for row in LOOKUP_ROWS}
unmatched_fields = sorted(
    f for f in lookup_fields
    if DESTINATION_FIELD_RENAMES.get(f, f) not in template_headers
)
 
print(f"\nBlock 5 done - {len(unmatched_fields)} Lookup field name(s) with no matching Template header:")
for f in unmatched_fields:
    print(f"  {f!r}")
 
# %%
# ============================================================
# BLOCK 6: WRITE REPORT
# ============================================================
 
with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["Field", "Sheet", "Source field (Lookup sheet)",
                                            "Status", "Resolved column"])
    writer.writeheader()
    writer.writerows(results)
 
print(f"\nBlock 6 done - full validation report written to {OUTPUT_FILE}")
 
# %%
 