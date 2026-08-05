# %% [markdown]
# # S62A migration - run this block by block
#
# Each `# %%` line below marks the start of a new "cell". In VS Code, you'll
# see a "Run Cell" / "Run Below" link appear above each one - click it (or
# put your cursor in the cell and press Shift+Enter) to run just that block.
#
# Variables stay alive between cells in the Interactive Window on the right,
# so after running BLOCK 4 you can type e.g. `df.head()` or `len(df)` at the
# bottom of the Interactive Window to poke at what you just loaded, before
# moving on to BLOCK 5.
#
# One-time setup if "Run Cell" doesn't appear: install the Jupyter extension
# in VS Code, then run:  pip3 install ipykernel

# %%
# ============================================================
# BLOCK 1: CONFIG
# ============================================================
import re
from datetime import datetime, date
import openpyxl
import pandas as pd

SOURCE_FILE = "Section-62a-Cases-COPY.xlsx"
TEMPLATE_FILE = "Template LEGACY cases S62A.xlsx"
TEMPLATE_SHEET = "Template"
TEMPLATE_FIRST_DATA_ROW = 3

OUTPUT_FILE = "output/S62A_All_Sheets_migrated.xlsx"
AUDIT_LOG_FILE = "output/migration_audit_log.csv"

print("Block 1 done - config set")

import os
for f in os.listdir():
    print(repr(f))

# %%
# ============================================================
# BLOCK 2a: MAPPING - Application (Major) only, to start with
# ============================================================
MAJOR_MAPPING = [
    ("Case reference",                      " Ref",                 "before_bracket", None),
    ("Application Status",                  None,                   "constant",       "Closed"),
    ("Pre-application or application",      None,                   "constant",       "Application"),
    ("Application classification",          None,                   "constant",       "Major"),
    ("Site address 1",                      "Address",               "address_part",   "address1"),
    ("Site address 2",                      "Address",               "address_part",   "address2"),
    ("Site town or city",                   "Address",               "address_part",   "town"),
    ("Site county",                         "Address",               "address_part",   "county"),
    ("Site post code",                      "Address",               "address_part",   "postcode"),
    ("LPA",                                 "LPA",                  "direct",         None),
    ("Likely issues",                       "Likely issues",        "direct",         None),
    ("Applicant organisation name",         "Applicant",            "direct",         None),
    ("Agent organisation name",             "Agent",                "agent_org",      None),
    ("Agent email",                         "Agent",                "agent_email",    None),
    ("Application type",                    "Type of application",  "direct",         None),
    ("Pre-app reference",                   "Pre-application reference", "direct",    None),
    ("Notification received date",          "Pre-notification date", "date",          None),
    ("Expected submission date",            "Expected submission date (not before)", "date", None),
    ("Pre-app/Application received date",   "Receipt date",          "date",          None),
    ("Development description",             "Development",          "direct",         None),
    ("Inspector 1",                         "Inspector",             "direct",         None),
    ("Date Inspector (1) allocated",        "Date Insp allocated",   "date",           None),
    ("Case officer",                        "Case Officer",          "direct",         None),
    ("Application fee amount",              "Fee amount",            "direct",         None),
    ("Application acknowledged",            "Letter 7 sent to applicant", "date",      None),
    ("LPA questionnaire sent",              "Letter 16 and Q sent to LPA", "date",     None),
    ("Application valid",                   "Valid Date",            "date",           None),
    ("Publish date",                        "Application documents published actual date", "date", None),
    ("LPA questionnaire rec'd date",        "LPA Q received",        "date",           None),
    ("CIL amount",                          None,                    "constant",       0),
    ("Representations period - start",      "Consultations sent",    "date",           None),
    ("Interim findings date",               "Inspectors Interim findings (letter 25 to applicant)", "date", None),
    ("Additional meeting required date",    "CMC date",              "date",           None),
    ("Hearing notification date",           "Hearing date notification (at least 2 weeks before hearing for major applications)", "date", None),
    ("Hearing issues report published date", "Issues report published date", "date",  None),
    ("S106 submitted date",                 "S106 submitted date",   "date",           None),
    ("Procedure",                           "Procedure WR/H?  Refer to PCU?", "direct", None),
    ("Hearing  date",                       "Hearing Date",          "date",           None),
    ("Target decision date",                "Target decision date (13 or 16 weeks from valid)", "date", None),
    ("Reader",                              "Reader",                "direct",         None),
    ("Decision date",                       "Decision date ",        "date",           None),
    ("Decision outcome",                    "Grant/Refuse",          "grant_refuse",   None),
    ("Notes",                               "Comments/Notes",        "direct",         None),
]

# Start with just this one sheet while you're testing - add the other two
# mapping lists (BLOCK 2b, 2c below) back into SHEET_CONFIGS once this works.
SHEET_CONFIGS = [
    ("Application (Major)", MAJOR_MAPPING),
]

print(f"Block 2 done - {len(MAJOR_MAPPING)} mapping rows defined for Major")


# %%
# ============================================================
# BLOCK 3: TRANSFORMS
# ============================================================
def is_blank(value):
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False

def transform_direct(value, extra):
    return None if is_blank(value) else value

def transform_date(value, extra):
    if is_blank(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    try:
        return pd.to_datetime(str(value).strip(), dayfirst=True).date()
    except (ValueError, TypeError):
        return None

def transform_constant(value, extra):
    return extra

def transform_before_bracket(value, extra):
    return None if is_blank(value) else str(value).split("(")[0].strip()

UK_POSTCODE_PATTERN = re.compile(r"\b([A-Za-z]{1,2}\d[A-Za-z\d]?\s*\d[A-Za-z]{2})\b")

def transform_postcode(value, extra):
    if is_blank(value):
        return None
    match = UK_POSTCODE_PATTERN.search(str(value))
    return match.group(1).upper() if match else None

def split_address_parts(text):
    """Best-effort split of one free-text address into its pieces.
    Returns a dict with keys: address1, address2, town, county, postcode."""
    if is_blank(text):
        return {}
    text = str(text).strip()

    postcode = None
    match = UK_POSTCODE_PATTERN.search(text)
    if match:
        postcode = match.group(1).upper()
        text = text[:match.start()].rstrip(", ")   # drop postcode + trailing comma

    parts = [p.strip() for p in text.split(",") if p.strip()]
    county = parts[-1] if len(parts) >= 1 else None
    town = parts[-2] if len(parts) >= 2 else None
    address_parts = parts[:-2] if len(parts) > 2 else []
    address1 = address_parts[0] if len(address_parts) >= 1 else None
    address2 = "; ".join(address_parts[1:]) if len(address_parts) > 1 else None

    return {"address1": address1, "address2": address2, "town": town,
            "county": county, "postcode": postcode}

def transform_address_part(value, extra):
    """extra = which piece to return: 'address1', 'address2', 'town', 'county', 'postcode'."""
    return split_address_parts(value).get(extra)

def transform_grant_refuse(value, extra):
    if is_blank(value):
        return None
    text = str(value).strip().lower()
    if "grant" in text:
        return "Granted"
    if "refus" in text:
        return "Refused"
    return None

def transform_labelled(value, extra):
    return None if is_blank(value) else f"{extra}: {value}"

EMAIL_PATTERN = re.compile(r"[\w.\-]+@[\w.\-]+")

def transform_agent_org(value, extra):
    """Takes the text before the first '(' as the org name, e.g. 'CBRE (a@x.com; b@y.com)' -> 'CBRE'."""
    if is_blank(value):
        return None
    return str(value).split("(")[0].strip()

def transform_agent_email(value, extra):
    """Pulls every email address out of the cell (there can be several), joined with '; '."""
    if is_blank(value):
        return None
    emails = EMAIL_PATTERN.findall(str(value))
    return "; ".join(emails) if emails else None

TRANSFORM_FUNCTIONS = {
    "direct": transform_direct, "date": transform_date, "constant": transform_constant,
    "before_bracket": transform_before_bracket, "postcode": transform_postcode,
    "grant_refuse": transform_grant_refuse, "labelled": transform_labelled,
    "address_part": transform_address_part,
    "agent_org": transform_agent_org, "agent_email": transform_agent_email,
}

print(f"Block 3 done - {len(TRANSFORM_FUNCTIONS)} transform functions ready")

# %%
# ============================================================
# BLOCK 4: READ SOURCE - just Application (Major) for now
# ============================================================
def read_source_rows(sheet_name):
    df = pd.read_excel(SOURCE_FILE, sheet_name=sheet_name, header=0)
    df = df.dropna(how="all")
    first_col = df.columns[0]
    df = df[df[first_col].notna()]
    return df

df_major = read_source_rows("Application (Major)")

print(f"Block 4 done - loaded {len(df_major)} rows")
print(df_major[[" Ref", "LPA", "Valid Date"]].head())
# ^ CHECK: do these look right? Wrong sheet name or column typo shows up here immediately.

# %%
# ============================================================
# BLOCK 5: BUILD ROWS - try it on ONE row first before all of them
# ============================================================
def build_output_rows(df, mapping, sheet_label):
    output_rows = []
    audit_entries = []
    for _, source_row in df.iterrows():
        row_result = {}
        for template_column, source_column, transform_name, extra in mapping:
            source_value = source_row.get(source_column) if source_column else None
            transform_fn = TRANSFORM_FUNCTIONS[transform_name]
            result = transform_fn(source_value, extra)
            if template_column in row_result and row_result[template_column] is not None and result is not None:
                row_result[template_column] = f"{row_result[template_column]}; {result}"
            elif result is not None or template_column not in row_result:
                row_result[template_column] = result
            if result is None and transform_name != "constant" and not is_blank(source_value):
                audit_entries.append((row_result.get("Case reference"), sheet_label, template_column,
                                       f"could not convert value: {source_value!r}"))
        output_rows.append(row_result)
    return output_rows, audit_entries

# --- try just the first row, so you can eyeball one full result ---
one_row_result, one_row_audit = build_output_rows(df_major.head(1), MAJOR_MAPPING, "Application (Major)")
print("Block 5 test done - single row result:")
for k, v in one_row_result[0].items():
    print(f"  {k}: {v!r}")
print(f"  ({len(one_row_audit)} audit flags for this row)")
# ^ CHECK: does every value look sensible? Fix the mapping/transform before running all rows.

# %%
# ============================================================
# BLOCK 5b: now run it on ALL rows
# ============================================================
all_rows, all_audit = build_output_rows(df_major, MAJOR_MAPPING, "Application (Major)")

print(f"Block 5b done - {len(all_rows)} rows built, {len(all_audit)} audit flags")

# %%
# ============================================================
# BLOCK 6: WRITE OUTPUT
# ============================================================
def write_output(all_rows):
    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb[TEMPLATE_SHEET]
    header_row = 2
    column_lookup = {}
    for col_num in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_num).value
        if header_value:
            column_lookup[header_value] = col_num
    for row_offset, row_result in enumerate(all_rows):
        excel_row = TEMPLATE_FIRST_DATA_ROW + row_offset
        for template_column, value in row_result.items():
            if template_column not in column_lookup:
                print(f"WARNING: '{template_column}' not found in Template headers - skipped")
                continue
            col_num = column_lookup[template_column]
            cell = ws.cell(row=excel_row, column=col_num)
            cell.value = value
            if isinstance(value, date):
                cell.number_format = "YYYY-MM-DD"
    import os
    os.makedirs(os.path.dirname(OUTPUT_FILE) or ".", exist_ok=True)
    wb.save(OUTPUT_FILE)

def write_audit_log(audit_entries):
    import csv, os
    os.makedirs(os.path.dirname(AUDIT_LOG_FILE) or ".", exist_ok=True)
    with open(AUDIT_LOG_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Case reference", "Source sheet", "Template column", "Issue"])
        writer.writerows(audit_entries)

write_output(all_rows)
write_audit_log(all_audit)

print(f"Block 6 done - wrote {len(all_rows)} rows to {OUTPUT_FILE}")
print(f"Audit log: {AUDIT_LOG_FILE} ({len(all_audit)} entries)")
# ^ CHECK: open the output file now and spot-check a few rows before adding
#   Non Major / Pre-application back into SHEET_CONFIGS in BLOCK 2.

# %%