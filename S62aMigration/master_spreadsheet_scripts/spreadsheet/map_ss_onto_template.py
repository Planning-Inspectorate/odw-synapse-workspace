import csv
import os
import re
from datetime import datetime, date
 
import openpyxl
import pandas as pd
from dateutil import parser as dateparser
 
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
 
def find_data_root(start_dir, marker="csv_and_xlsx_files", max_up=4):
    """The script has moved folders more than once (it now lives one level
    deeper, in .../master_spreadsheet_scripts/spreadsheet/), so rather than
    hardcode how many '..' to use, walk upward until the data folder is
    found. Works whether the script sits next to csv_and_xlsx_files or one
    or two levels below it."""
    d = start_dir
    for _ in range(max_up + 1):
        if os.path.isdir(os.path.join(d, marker)):
            return d
        parent = os.path.dirname(d)
        if parent == d:
            break
        d = parent
    raise FileNotFoundError(
        f"Could not find a '{marker}' folder above {start_dir} "
        f"(searched {max_up + 1} levels up) - check the script's location."
    )
 
DATA_ROOT = find_data_root(BASE_DIR)
 
SOURCE_FILE   = os.path.join(DATA_ROOT, "csv_and_xlsx_files/SS_data/Section-62a-Cases-COPY.xlsx")
MAPPING_XLSX  = os.path.join(DATA_ROOT, "csv_and_xlsx_files/SS_data/S62A_Column_mapping.xlsx")
MAPPING_SHEET = "Lookup"
TEMPLATE_FILE = os.path.join(DATA_ROOT, "csv_and_xlsx_files/MASTER LEGACY cases S62A .xlsx")
TEMPLATE_SHEET          = "Template"
TEMPLATE_HEADER_ROW     = 2
TEMPLATE_FIRST_DATA_ROW = 3
 
OUTPUT_FILE          = os.path.join(DATA_ROOT, "outputs/S62A_All_Sheets_migrated.xlsx")
AUDIT_LOG_FILE        = os.path.join(DATA_ROOT, "outputs/spreadsheet_migration_audit_log.csv")
UNMAPPED_COLS_FILE    = os.path.join(DATA_ROOT, "outputs/spreadsheet_unmapped_columns_report.csv")
MAPPING_ISSUES_FILE   = os.path.join(DATA_ROOT, "outputs/spreadsheet_mapping_config_issues.csv")
 
os.makedirs(os.path.join(DATA_ROOT, "outputs"), exist_ok=True)
 
print("Block 1 done - config set")
 
# %%
# ============================================================
# BLOCK 2: LOAD MAPPING - from the Lookup tab of S62A_Column_mapping.xlsx
# ============================================================
# Lookup sheet layout (row 1 = headers, row 2 onwards = data):
#   A: Category (sparse - only set on the first row of each group)
#   B: Field (final Template column name)
#   C: Pre-application - DONE  -> source column name for that sheet, or "N/A",
#                                  or literal text like Set to be "Major"
#   D: Application (Major)     -> same idea
#   E: Application (Non Major) -> same idea
#   F: Horizon                 -> not used here (Nisali's script handles Horizon)
#   G: Notes
 
SET_TO_BE_RE = re.compile(r'set to be\s+"([^"]+)"', re.IGNORECASE)
 
def load_lookup_mapping():
    wb = openpyxl.load_workbook(MAPPING_XLSX, data_only=True)
    ws = wb[MAPPING_SHEET]
    rows = []
    current_category = None
    for r in range(2, ws.max_row + 1):
        category = ws.cell(row=r, column=1).value
        field    = ws.cell(row=r, column=2).value
        if category:
            current_category = category
        if not field:
            continue
        rows.append({
            "category": current_category,
            "field": str(field).strip(),
            "Pre-application - DONE":  ws.cell(row=r, column=3).value,
            "Application (Major)":     ws.cell(row=r, column=4).value,
            "Application (Non Major)": ws.cell(row=r, column=5).value,
            "notes": ws.cell(row=r, column=7).value,
        })
    return rows
 
LOOKUP_ROWS = load_lookup_mapping()
 
# Baseline constants that apply regardless of what the Lookup sheet says,
# because they're known project facts rather than per-field mappings.
# (All S62A Horizon/spreadsheet cases are closed - see project memory.)
SHEET_CONSTANTS = {
    "Pre-application - DONE":  {"Application Status": "Closed"},
    "Application (Major)":     {"Application Status": "Closed"},
    "Application (Non Major)": {"Application Status": "Closed"},
}
 
# --- Fields the Lookup sheet marks "N/A" (no direct source column) but whose
# Notes say they should be pulled out of another field. Only added here where
# the split is actually reliable from the real data - see the message this
# script was delivered with for the ones deliberately left out.
MANUAL_SOURCE_OVERRIDES = {
    "Site address 2":    ("Address", "address_part", "address2"),
    "Site town or city":  ("Address", "address_part", "town"),
    "Site county":        ("Address", "address_part", "county"),
    "Site post code":     ("Address", "address_part", "postcode"),
    "Agent email":        ("Agent",   "agent_email",  None),
}
 
# --- Transform overrides for fields that need something other than a
# straight copy or the default date/direct guess.
FIELD_TRANSFORM_OVERRIDES = {
    "Case reference":          ("before_bracket", None),
    "Site address 1":          ("address_part", "address1"),
    "Agent organisation name": ("agent_org", None),
    "Decision outcome":        ("grant_refuse", None),
    "Site visit type":         ("site_visit_type", None),
    "Site visit date":         ("site_visit_date", None),
    "CIL liable":               ("cil_liable", None),
    "CIL amount":                ("cil_amount", None),
    "Inspector band":          ("specialism_band", "band"),
    "Specialism":              ("specialism_band", "specialism"),
    "Press notice placed":     ("press_notice", "placed"),
    "Press notice reference":  ("press_notice", "reference"),
    "EIA screening outcome":              ("eia_outcome", None),
    "Date environment statement was received": ("eia_received_date", None),
    "Date Environmental Statement rec'd":      ("eia_received_date", None),
    "Customer number":         ("customer_number", None),
    "Pre-application fee":     ("fee_amount", None),
    "Pre-application fee due": ("fee_amount", None),
}
 
# The Major sheet has no reliable numeric CIL figure in the source data -
# "LPA CIL response" is free text and is "NA" for every Major row we checked,
# and there's no separate "CIL amount" column on that sheet at all. Leave it
# unmapped there rather than pointing "CIL amount" at a column that can't
# produce a number. (Non-major has a real "CIL amount" column, so it's fine.)
SHEET_FIELD_SKIPS = {
    "Application (Major)": {"CIL amount"},
}
 
# The Lookup sheet's source-column text and the real spreadsheet headers
# disagree on whitespace/case in several places (' Ref' vs 'Ref', 'Decision
# date ' vs 'Decision date', 'Likely Issues' vs 'Likely issues', etc) - those
# are resolved automatically below. A couple of columns were genuinely
# renamed on one sheet and need an explicit alias instead.
SOURCE_COLUMN_ALIASES = {
    "Application (Non Major)": {
        "SAP5 to FSSD": "SAP5 to FSSD / Fee requested by BACS",
    },
}
 
def _norm_col(text):
    return re.sub(r"\s+", " ", str(text).strip()).lower()
 
def resolve_source_column(sheet_name, source_col, available_by_norm, available):
    """Match a Lookup-sheet source column name against the real sheet headers,
    tolerating whitespace/case differences and known renames."""
    alias = SOURCE_COLUMN_ALIASES.get(sheet_name, {}).get(source_col)
    if alias:
        source_col = alias
    if source_col in available:
        return source_col
    return available_by_norm.get(_norm_col(source_col))
 
def build_mapping_for_sheet(sheet_name, available_columns):
    """Returns (mapping, constants, config_issues) for one source sheet.
    mapping = list of (template_field, source_column, transform_name, extra)
    constants = {template_field: value} to apply to every row on this sheet
    config_issues = Lookup rows pointing at a column that doesn't exist here
    """
    mapping = []
    config_issues = []
    available = set(available_columns)
    available_by_norm = {_norm_col(c): c for c in available_columns}
    constants = dict(SHEET_CONSTANTS.get(sheet_name, {}))
 
    for row in LOOKUP_ROWS:
        field = row["field"]
        if field in SHEET_FIELD_SKIPS.get(sheet_name, set()):
            continue
 
        if field in MANUAL_SOURCE_OVERRIDES:
            source_col, transform_name, extra = MANUAL_SOURCE_OVERRIDES[field]
        else:
            raw_source = row.get(sheet_name)
            raw_source = str(raw_source).strip() if raw_source is not None else ""
            if not raw_source or raw_source.upper() == "N/A":
                continue
            m = SET_TO_BE_RE.search(raw_source)
            if m:
                constants[field] = m.group(1)
                continue
            source_col = raw_source
            transform_name, extra = FIELD_TRANSFORM_OVERRIDES.get(field, (None, None))
            if transform_name is None:
                transform_name, extra = ("date", None) if "date" in field.lower() else ("direct", None)
 
        resolved_col = resolve_source_column(sheet_name, source_col, available_by_norm, available)
        if resolved_col is None:
            config_issues.append((sheet_name, field, source_col,
                                   "source column not found on this sheet"))
            continue
 
        mapping.append((field, resolved_col, transform_name, extra))
 
    return mapping, constants, config_issues
 
print(f"Block 2 done - {len(LOOKUP_ROWS)} template fields loaded from Lookup sheet")
 
# %%
# ============================================================
# BLOCK 3: TRANSFORMS
# ============================================================
def is_blank(value):
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False
 
def transform_direct(value, extra):
    return None if is_blank(value) else value
 
def transform_date(value, extra):
    if is_blank(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    try:
        return pd.to_datetime(str(value).strip(), dayfirst=True).date()
    except (ValueError, TypeError):
        return None
 
def transform_before_bracket(value, extra):
    return None if is_blank(value) else str(value).split("(")[0].strip()
 
UK_POSTCODE_PATTERN = re.compile(r"\b([A-Za-z]{1,2}\d[A-Za-z\d]?\s*\d[A-Za-z]{2})\b")
 
def split_address_parts(text):
    """Best-effort split of one free-text address into its pieces."""
    if is_blank(text):
        return {}
    text = str(text).strip()
 
    postcode = None
    match = UK_POSTCODE_PATTERN.search(text)
    if match:
        postcode = match.group(1).upper()
        text = text[:match.start()].rstrip(", ")
 
    parts = [p.strip() for p in text.split(",") if p.strip()]
    county = parts[-1] if len(parts) >= 1 else None
    town = parts[-2] if len(parts) >= 2 else None
    address_parts = parts[:-2] if len(parts) > 2 else []
    address1 = address_parts[0] if len(address_parts) >= 1 else None
    address2 = "; ".join(address_parts[1:]) if len(address_parts) > 1 else None
 
    return {"address1": address1, "address2": address2, "town": town,
            "county": county, "postcode": postcode}
 
def transform_address_part(value, extra):
    return split_address_parts(value).get(extra)
 
def transform_grant_refuse(value, extra):
    if is_blank(value):
        return None
    text = str(value).strip().lower()
    if "grant" in text:
        return "Granted"
    if "refus" in text:
        return "Refused"
    return None
 
EMAIL_PATTERN = re.compile(r"[\w.\-]+@[\w.\-]+")
 
def transform_agent_org(value, extra):
    """Org name before the '(' when there's a clean 'Org (emails)' pattern.
    Otherwise (name-only, bare email, 'Name <email>', etc) there's nothing
    reliable to split, so the whole raw value is kept here rather than
    guessing - per Abbas: keep all emails in Agent email, and if the Agent
    data doesn't split cleanly just keep everything together in one field."""
    if is_blank(value):
        return None
    return str(value).split("(")[0].strip()
 
def transform_agent_email(value, extra):
    if is_blank(value):
        return None
    emails = EMAIL_PATTERN.findall(str(value))
    return "; ".join(emails) if emails else None
 
# --- Level & Specialism, e.g. "B3 RE" -> Inspector band "Band 3", Specialism
# "RE". Not every sheet has the band prefix (Non-major values like "HG" have
# no band at all) - in that case the whole value is kept as Specialism and
# Inspector band is left blank, rather than guessing a band.
BAND_TOKEN_RE = re.compile(r'^B(?:AND)?\s*(\d+)$', re.IGNORECASE)
 
def split_specialism_band(text):
    if is_blank(text):
        return {}
    text = str(text).strip()
    tokens = text.split(None, 1)
    band = None
    specialism = text
    if tokens:
        m = BAND_TOKEN_RE.match(tokens[0])
        if m:
            band = f"Band {m.group(1)}"
            specialism = tokens[1].strip() if len(tokens) > 1 else None
    return {"band": band, "specialism": specialism}
 
def transform_specialism_band(value, extra):
    return split_specialism_band(value).get(extra)
 
# --- Press notice: one free-text cell holds a placed-date plus a reference/
# cost note, e.g. "16 June 2022.  Saffron Walden Reporter...  £337.10 plus
# VAT" or "27-Oct-22 (£497.54) TMP ref 443455". Per Abbas: the first part
# looks like the date, so pull that off the front as "placed" and put
# everything else into "reference". If the text doesn't start with a
# recognisable date, leave placed blank and keep the whole value in
# reference rather than losing anything.
PRESS_NOTICE_DATE_RE = re.compile(
    r'^\s*(\d{1,2}\s*[/\-]\s*\d{1,2}\s*[/\-]\s*\d{2,4}'   # 23/2/23, 27-10-22
    r'|\d{1,2}\s+[A-Za-z]+\s+\d{2,4}'                      # 16 June 2022
    r'|\d{1,2}\s*-\s*[A-Za-z]+\s*-\s*\d{2,4})'             # 27-Oct-22
)
 
def split_press_notice(value):
    if is_blank(value):
        return {}
    if isinstance(value, (datetime, date)):
        return {"placed": value.date() if isinstance(value, datetime) else value,
                "reference": None}
    text = str(value).strip()
    match = PRESS_NOTICE_DATE_RE.match(text)
    if not match:
        return {"placed": None, "reference": text}
    leading = match.group(1)
    remainder = text[match.end():].strip(" .")
    try:
        placed = dateparser.parse(leading, fuzzy=True, dayfirst=True).date()
    except (ValueError, TypeError, OverflowError):
        return {"placed": None, "reference": text}
    return {"placed": placed, "reference": remainder or None}
 
def transform_press_notice(value, extra):
    return split_press_notice(value).get(extra)
 
GBP_AMOUNT_RE = re.compile(r'£\s?([\d,]+(?:\.\d+)?)')
 
def extract_gbp_amount(text):
    if not text:
        return None
    m = GBP_AMOUNT_RE.search(str(text))
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except ValueError:
        return None
 
# --- EIA outcome / received date: both fields share one free-text source
# column ("EIA outcome (publish any letter to gov.uk)"). In the real data
# this column is always a descriptive sentence, not a bare date, so the
# outcome text is kept as-is; the received date is only pulled out when the
# text actually mentions submission/receipt near a date (e.g. "ES requested
# 5/10/22. Submtted 23/12/22") - otherwise it's left blank rather than
# misreading an unrelated date (like the screening-request date) as the
# received date.
def transform_eia_outcome(value, extra):
    return transform_direct(value, extra)
 
EIA_RECEIVED_RE = re.compile(
    r"(?:submit|receiv|rec'?d)[a-z]*\D{0,15}?"
    r"(\d{1,2}\s*[/\-]\s*\d{1,2}\s*[/\-]\s*\d{2,4}|\d{1,2}\s+[A-Za-z]+\s+\d{2,4})",
    re.IGNORECASE,
)
 
def transform_eia_received_date(value, extra):
    if is_blank(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    match = EIA_RECEIVED_RE.search(str(value))
    if not match:
        return None
    try:
        return dateparser.parse(match.group(1), fuzzy=True, dayfirst=True).date()
    except (ValueError, TypeError, OverflowError):
        return None
 
# --- Customer number: pulled out of the free-text "Set up as customer for
# invoicing? (SAP 8)" column, e.g. "Customer number set up 708832" or
# "15/03/2024. Customer no 710585". PINS customer numbers in this data are
# consistently 6 digits.
CUSTOMER_NUMBER_RE = re.compile(r'\b(\d{6})\b')
 
def transform_customer_number(value, extra):
    if is_blank(value):
        return None
    match = CUSTOMER_NUMBER_RE.search(str(value))
    return match.group(1) if match else None
 
# --- Fee amount: mostly plain numbers but occasionally free text like
# "£6951 plus mileage for SV" - pull the leading number out and drop the
# rest rather than losing the whole value.
def transform_fee_amount(value, extra):
    if is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return value
    match = re.search(r'[\d,]+(?:\.\d+)?', str(value).replace('£', ''))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None
 
# --- Site visit: one free-text/date cell needs to become type (USV/ARSV) and
# date separately. Source data is inconsistent (plain dates, "USV 8 October
# 2024", "ARSV 6 May 2025 3.00-4.00pm", "USV w/c 24/03/2025", and the odd
# unrelated note like "See applicant email of ..."), so both of these are
# best-effort - check the audit log for rows where the date couldn't be
# parsed at all.
SITE_VISIT_TYPE_RE = re.compile(r"\b(ARSV|USV)\b", re.IGNORECASE)
 
def transform_site_visit_type(value, extra):
    if is_blank(value):
        return None
    if isinstance(value, (datetime, date)):
        return None  # plain date with no type mentioned
    match = SITE_VISIT_TYPE_RE.search(str(value))
    return match.group(1).upper() if match else None
 
def transform_site_visit_date(value, extra):
    if is_blank(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    try:
        return dateparser.parse(str(value), fuzzy=True, dayfirst=True).date()
    except (ValueError, TypeError, OverflowError):
        return None
 
# --- CIL liable: "LPA CIL response" is free text like "Not CIL liable
# 26/09/2025", "CIL liable 8/10/24", "NA", "??". Treat "not"/"no" + liable
# as No, "liable" on its own as Yes, anything else (NA, ??, blank) as
# unknown/blank rather than guessing.
def transform_cil_liable(value, extra):
    if is_blank(value):
        return None
    text = str(value).strip().lower()
    if text in ("na", "n/a", "??"):
        return None
    if "not" in text or "no cil" in text or re.search(r"\bnot\b", text):
        return "No"
    if "liable" in text:
        return "Yes"
    return None
 
def transform_cil_amount(value, extra):
    if is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text.upper() in ("NA", "N/A"):
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None
 
TRANSFORM_FUNCTIONS = {
    "direct": transform_direct,
    "date": transform_date,
    "before_bracket": transform_before_bracket,
    "address_part": transform_address_part,
    "grant_refuse": transform_grant_refuse,
    "agent_org": transform_agent_org,
    "agent_email": transform_agent_email,
    "site_visit_type": transform_site_visit_type,
    "site_visit_date": transform_site_visit_date,
    "cil_liable": transform_cil_liable,
    "cil_amount": transform_cil_amount,
    "specialism_band": transform_specialism_band,
    "press_notice": transform_press_notice,
    "eia_outcome": transform_eia_outcome,
    "eia_received_date": transform_eia_received_date,
    "customer_number": transform_customer_number,
    "fee_amount": transform_fee_amount,
}
 
print(f"Block 3 done - {len(TRANSFORM_FUNCTIONS)} transform functions ready")
 
# %%
# ============================================================
# BLOCK 4: READ SOURCE - just the first sheet for now
# ============================================================
SHEET_NAMES = ["Pre-application - DONE", "Application (Major)", "Application (Non Major)"]
 
def read_source_rows(sheet_name):
    df = pd.read_excel(SOURCE_FILE, sheet_name=sheet_name, header=0)
    df = df.dropna(how="all")
    first_col = df.columns[0]
    df = df[df[first_col].notna()]
    return df
 
_test_sheet = SHEET_NAMES[1]  # Application (Major) - has the richest data
df_test = read_source_rows(_test_sheet)
 
print(f"Block 4 done - {_test_sheet}: loaded {len(df_test)} rows")
print(df_test.head())
 
# %%
# ============================================================
# BLOCK 5: BUILD ROWS - try it on ONE row first before all of them
# ============================================================
def build_output_rows(df, mapping, constants, sheet_label):
    output_rows   = []
    audit_entries = []
    for _, source_row in df.iterrows():
        row_result = dict(constants)
        for template_column, source_column, transform_name, extra in mapping:
            source_value = source_row.get(source_column)
            transform_fn = TRANSFORM_FUNCTIONS[transform_name]
            result = transform_fn(source_value, extra)
            if result is not None:
                existing = row_result.get(template_column)
                if existing and str(result) not in str(existing).split("; "):
                    row_result[template_column] = f"{existing}; {result}"
                elif not existing:
                    row_result[template_column] = result
            elif transform_name not in ("address_part", "site_visit_type",
                                         "specialism_band", "press_notice",
                                         "eia_received_date", "customer_number") \
                    and not is_blank(source_value):
                # site_visit_type/specialism_band/press_notice/eia_received_date/
                # customer_number each have a legitimately-blank outcome quite
                # often (no USV/ARSV keyword, no band prefix, no leading date,
                # no receipt date mentioned, no 6-digit number found) - that's
                # a normal outcome of the free text, not a real conversion
                # failure, so it's excluded from the audit log noise.
                audit_entries.append((row_result.get("Case reference"), sheet_label,
                                       template_column, f"could not convert value: {source_value!r}"))
        # Press notice cost cross-check: the direct "TMP amount" column is
        # usually enough, but where it's blank the cost is often still sat in
        # the press notice free text, e.g. "(£497.54) TMP ref 443455" - which
        # by this point has already landed in Press notice reference above.
        if not row_result.get("Press notice cost"):
            cost = extract_gbp_amount(row_result.get("Press notice reference"))
            if cost is not None:
                row_result["Press notice cost"] = cost
        output_rows.append(row_result)
    return output_rows, audit_entries
 
mapping_test, constants_test, issues_test = build_mapping_for_sheet(_test_sheet, df_test.columns.tolist())
one_row_result, one_row_audit = build_output_rows(df_test.head(1), mapping_test, constants_test, _test_sheet)
 
print(f"Block 5 test done - {_test_sheet}: {len(mapping_test)} mapped columns, {len(issues_test)} config issues")
print("Single row result:")
for k, v in one_row_result[0].items():
    print(f"  {k}: {v!r}")
print(f"  ({len(one_row_audit)} audit flags for this row)")
# ^ CHECK: does every value look sensible? Fix FIELD_TRANSFORM_OVERRIDES/MANUAL_SOURCE_OVERRIDES before continuing.
 
# %%
# ============================================================
# BLOCK 5b: now run it on ALL rows, for every sheet
# ============================================================
all_rows        = []
all_audit       = []
all_config_issues = []
unmapped_report = []  # {Sheet, Column} - source columns no template field claims
 
for sheet_name in SHEET_NAMES:
    df = read_source_rows(sheet_name)
    mapping, constants, config_issues = build_mapping_for_sheet(sheet_name, df.columns.tolist())
    all_config_issues.extend(config_issues)
 
    used_columns = {source_col for _, source_col, _, _ in mapping}
    for col in df.columns:
        if col not in used_columns:
            unmapped_report.append({"Sheet": sheet_name, "Column": col})
 
    rows, audit = build_output_rows(df, mapping, constants, sheet_name)
    all_rows.extend(rows)
    all_audit.extend(audit)
    print(f"  {sheet_name}: {len(mapping)} mapped columns, {len(config_issues)} config issues, {len(df)} rows")
 
print(f"Block 5b done - {len(all_rows)} rows built, {len(all_audit)} audit flags, {len(all_config_issues)} config issues")
 
# %%
# ============================================================
# BLOCK 6: WRITE OUTPUT
# ============================================================
# A few field names in the Lookup sheet don't match the real Template
# header - the Lookup sheet uses "Pre-application or application" but the
# Template column is actually called "Application phase" (its value is
# still just the pre-app/major/non-major constant, telling us which of the
# three source sheets a row came from).
DESTINATION_FIELD_RENAMES = {
    "Pre-application or application": "Application phase",
}
 
os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
wb = openpyxl.load_workbook(TEMPLATE_FILE)
ws = wb[TEMPLATE_SHEET]
 
column_lookup = {}
for col_num in range(1, ws.max_column + 1):
    header_value = ws.cell(row=TEMPLATE_HEADER_ROW, column=col_num).value
    if header_value:
        column_lookup[str(header_value).strip()] = col_num
 
_missing_fields = set()
excel_row = TEMPLATE_FIRST_DATA_ROW
for row_result in all_rows:
    for col_num in column_lookup.values():
        ws.cell(row=excel_row, column=col_num).value = None
    for template_column, value in row_result.items():
        template_column = DESTINATION_FIELD_RENAMES.get(template_column, template_column)
        if template_column not in column_lookup:
            _missing_fields.add(template_column)
            print(f"WARNING: '{template_column}' not found in Template headers - skipped")
            continue
        col_num = column_lookup[template_column]
        cell = ws.cell(row=excel_row, column=col_num)
        cell.value = value
        if isinstance(value, date):
            cell.number_format = "YYYY-MM-DD"
    excel_row += 1
 
wb.save(OUTPUT_FILE)
print(f"Block 6 done - wrote {len(all_rows)} rows to {OUTPUT_FILE}")
 
if _missing_fields:
    print(f"\n{len(_missing_fields)} unmatched Field name(s) - compare against real Template headers below")
    print("Unmatched Field names from the Lookup sheet:")
    for f in sorted(_missing_fields):
        print(f"  {f!r}")
    print("Real Template headers:")
    for h in sorted(column_lookup):
        print(f"  {h!r}")
 
# %%
# ============================================================
# BLOCK 7: AUDIT LOG + UNMAPPED COLUMNS + MAPPING CONFIG ISSUES
# ============================================================
os.makedirs(os.path.dirname(AUDIT_LOG_FILE), exist_ok=True)
 
with open(AUDIT_LOG_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Case reference", "Source sheet", "Template column", "Issue"])
    writer.writerows(all_audit)
print(f"Audit log:               {AUDIT_LOG_FILE} ({len(all_audit)} entries)")
 
with open(UNMAPPED_COLS_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["Sheet", "Column"])
    writer.writeheader()
    writer.writerows(unmapped_report)
print(f"Unmapped columns report:  {UNMAPPED_COLS_FILE} ({len(unmapped_report)} entries)")
 
with open(MAPPING_ISSUES_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Sheet", "Field", "Source column (from Lookup sheet)", "Issue"])
    writer.writerows(all_config_issues)
print(f"Mapping config issues:    {MAPPING_ISSUES_FILE} ({len(all_config_issues)} entries)")
 
# %%
 