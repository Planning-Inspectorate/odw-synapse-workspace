# # Map all Horizon extracts onto MASTER LEGACY cases S62A
#
# Follows the same block-by-block pattern as ssmapping_stepbystep.py.
#
# horizon_field_mapping.csv tells us:
#   'Source field' = extract filename token and column name
#   'Field'        = column name in MASTER LEGACY cases S62A .xlsx
#   'Conditions'   = supported WHERE filters and special transformations
#
# each CSV in Horizon_extracts/ is processed in turn. Because multiple CSVs
# share the same CaseReference (each file just has different columns for the
# same case), rows from all files are merged into ONE row per case in the
# output. Any CSV columns that have no entry in horizon_field_mapping.csv are
# flagged in a separate report.

# CONFIG
import csv
import os
import re
import shutil
from datetime import datetime, date

import openpyxl
import pandas as pd

BASE_DIR = "/Users/nisalihalwathura/PINS/ODW-Service/odw-synapse-workspace/S62aMigration"

HORIZON_EXTRACTS_DIR = os.path.join(BASE_DIR, "csv_and_xlsx_files/Horizon_extracts")
MAPPING_CSV          = os.path.join(BASE_DIR, "outputs/horizon_field_mapping.csv")
MASTER_FILE          = os.path.join(BASE_DIR, "csv_and_xlsx_files/MASTER LEGACY cases S62A .xlsx")
TEMPLATE_SHEET       = "Template"
TEMPLATE_HEADER_ROW     = 2
TEMPLATE_FIRST_DATA_ROW = 3

OUTPUT_FILE          = os.path.join(BASE_DIR, "outputs/MASTER LEGACY cases S62A - with Horizon data.xlsx")
AUDIT_LOG_FILE       = os.path.join(BASE_DIR, "outputs/horizon_migration_audit_log.csv")
UNMAPPED_COLS_FILE   = os.path.join(BASE_DIR, "outputs/horizon_unmapped_columns_report.csv")

print("Block 1 done - config set")


# MAPPING HELPER
# Source fields identify an extract by the token before the first dot. A token
# can match more than one export, so files are resolved separately.

_camel      = re.compile(r"(?<=[a-z0-9])(?=[A-Z])")
def _norm(t): return re.sub(r"\s+", " ", _camel.sub(" ", str(t)).lower()).strip()

_mapping_df = pd.read_csv(MAPPING_CSV, dtype=str).fillna("")

def _source_parts(source_field):
    """Return (filename token, source columns) from a Source field value."""
    token, separator, columns = source_field.partition(".")
    if not separator:
        return "", []
    source_columns = []
    for part in columns.split(","):
        part = part.strip()
        if "." in part:
            part = part.rsplit(".", 1)[-1].strip()
        if part and part not in source_columns:
            source_columns.append(part)
    return token.strip(), source_columns

def _condition_filter(condition):
    """Return (condition column, accepted values) for a supported WHERE rule."""
    if not condition.strip().upper().startswith("WHERE"):
        return None
    match = re.search(
        r"WHERE\s+[^.\s]+\.([\w]+)\s*(?:=\s*[\"']([^\"']+)[\"']|in\s*\(([^)]*)\))",
        condition,
        re.IGNORECASE,
    )
    if not match:
        return None
    values = match.group(2)
    if values is None:
        values = [value.strip().strip("\"'") for value in match.group(3).split(",")]
    else:
        values = [values.strip()]
    return match.group(1), {value.casefold() for value in values}

_mapping_rules = []
for _, _r in _mapping_df.iterrows():
    _field = _r["Field"].strip()
    _token, _columns = _source_parts(_r["Source field"].strip())
    if _field and _token and _columns:
        _mapping_rules.append({
            "template_column": _field,
            "file_token": _token,
            "source_columns": _columns,
            "condition": _condition_filter(_r["Conditions"]),
            "append": "append to array" in _r["Conditions"].strip().casefold(),
            "aggregate_by_case": _token.casefold() == "extended_data",
        })

_extended_data_columns = {
    rule["template_column"]
    for rule in _mapping_rules
    if rule["aggregate_by_case"]
}

def build_mapping_for_file(filepath):
    """Returns (MAPPING, unmapped_columns) for the given CSV file.
    MAPPING  = list of mapping rule dictionaries
    unmapped = list of column names that have no entry in horizon_field_mapping
    """
    headers = pd.read_csv(filepath, nrows=0).columns.tolist()
    filename = os.path.basename(filepath)
    mapping = []
    mapped_columns = set()
    for rule in _mapping_rules:
        if rule["file_token"] not in filename:
            continue
        available = [column for column in rule["source_columns"] if column in headers]
        if available:
            mapping.append({**rule, "source_columns": available})
            mapped_columns.update(available)
    unmapped = [col for col in headers if col not in mapped_columns and col not in {"CaseReference", "CaseUniqueId", "CaseNodeId"}]
    return mapping, unmapped

print("Block 2 done - mapping helper ready")


# TRANSFORMS
def is_blank(value):
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() in ("", "NULL", "null"):
        return True
    return False

def transform_direct(value, extra):
    return None if is_blank(value) else str(value).strip()

def transform_email(value):
    value = transform_direct(value, None)
    return None if value is None or re.fullmatch(r"\d+(?:\.\d+)?", value) else value

def transform_date(value, extra):
    if is_blank(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    try:
        return pd.to_datetime(str(value).strip(), dayfirst=True).date()
    except (ValueError, TypeError):
        return None

def transform_constant(value, extra):
    return extra

def transform_hearing_duration(source_row):
    end_value = source_row.get("EndDate")
    start_value = source_row.get("StartDate")
    if is_blank(end_value) or is_blank(start_value):
        return None
    try:
        end_date = pd.to_datetime(str(end_value).strip(), dayfirst=True)
        start_date = pd.to_datetime(str(start_value).strip(), dayfirst=True)
        if pd.isna(end_date) or pd.isna(start_date):
            raise ValueError
        return (end_date - start_date).days
    except (ValueError, TypeError):
        return f"{str(end_value).strip()} - {str(start_value).strip()}"

TRANSFORM_FUNCTIONS = {
    "direct":   transform_direct,
    "date":     transform_date,
    "constant": transform_constant,
}

print(f"Block 3 done - {len(TRANSFORM_FUNCTIONS)} transform functions ready")


# LOOP ALL CSVs AND BUILD ONE ROW PER CASE
# case_rows  : {CaseReference: {template_column: value}}  - merged across all files
# all_audit  : list of (case_ref, template_col, issue)
# unmapped_report: list of {file, column} for columns not in the mapping CSV

def build_output_rows(df, mapping):
    output_rows   = []
    audit_entries = []
    for _, source_row in df.iterrows():
        row_result = {}
        if "CaseReference" in source_row and not is_blank(source_row["CaseReference"]):
            row_result["Case reference"] = str(source_row["CaseReference"]).strip()
        for rule in mapping:
            template_column = rule["template_column"]
            if template_column == "Case reference":
                continue
            condition = rule["condition"]
            if condition:
                condition_column, accepted_values = condition
                condition_value = source_row.get(condition_column)
                if is_blank(condition_value) or str(condition_value).strip().casefold() not in accepted_values:
                    continue

            if template_column == "Hearing duration - sitting (days)":
                source_value = transform_hearing_duration(source_row)
            elif template_column == "Application phase":
                category = str(source_row.get("caseCategory", "")).strip().casefold()
                source_value = "pre-application" if category == "pre-app" else "application"
            elif template_column == "Application subtype":
                category = str(source_row.get("caseCategory", "")).strip().casefold()
                source_value = "Listed building consent" if category == "listed-building" else "Planning permission"
            else:
                values = [source_row.get(column) for column in rule["source_columns"]]
                source_value = ", ".join(str(value).strip() for value in values if not is_blank(value))
            result = transform_email(source_value) if template_column == "LPA contact - Email" else transform_direct(source_value, None)
            if result is not None:
                if template_column in row_result and row_result[template_column] is not None:
                    row_result[template_column] = f"{row_result[template_column]}; {result}"
                else:
                    row_result[template_column] = result
            elif not is_blank(source_value):
                audit_entries.append((
                    row_result.get("Case reference"),
                    template_column,
                    f"could not convert value: {source_value!r}"
                ))
        output_rows.append(row_result)
    return output_rows, audit_entries

# CaseReference -> each record contains values, conflict cells, and
# unique values for fields whose mapping says to append to an array.
case_rows      = {}
all_audit      = []
unmapped_report = []  # {file, column}

# Load the authoritative list of S62A case references so we only process
# rows that belong to S62A cases (some CSVs like case_involvement_o cover
# the entire Horizon system and would otherwise bring in hundreds of thousands
# of non-S62A cases).
_ref_file = os.path.join(HORIZON_EXTRACTS_DIR, "20260716_query_s62A_case_reference_list.csv")
_reference_df = pd.read_csv(_ref_file, dtype=str).fillna("")
_reference_df["CaseUniqueId"] = _reference_df["CaseUniqueId"].str.strip()
_reference_df["CaseReference"] = _reference_df["CaseReference"].str.strip()
CASE_UNIQUE_ID_TO_REFERENCE = dict(
    zip(_reference_df["CaseUniqueId"], _reference_df["CaseReference"])
)
S62A_CASES = set(_reference_df["CaseReference"]) - {""}
print(f"  S62A case reference list loaded: {len(S62A_CASES)} cases")

csv_files = sorted(f for f in os.listdir(HORIZON_EXTRACTS_DIR) if f.endswith(".csv"))

for filename in csv_files:
    filepath = os.path.join(HORIZON_EXTRACTS_DIR, filename)
    mapping, unmapped_cols = build_mapping_for_file(filepath)

    # Record unmapped columns for this file
    for col in unmapped_cols:
        unmapped_report.append({"File": filename, "Column": col})

    if not mapping:
        print(f"  {filename}: no mapped columns - skipped")
        continue

    df = pd.read_csv(filepath, dtype=str).dropna(how="all")
    if "CaseReference" not in df.columns and "CaseUniqueId" in df.columns:
        df["CaseReference"] = (
            df["CaseUniqueId"].fillna("").str.strip().map(CASE_UNIQUE_ID_TO_REFERENCE)
        )
    if "CaseReference" in df.columns:
        df = df[df["CaseReference"].str.strip().isin(S62A_CASES)]
    rows, audit = build_output_rows(df, mapping)
    all_audit.extend(audit)

    # Merge into case_rows by CaseReference; extended_data fields are combined
    # unique values, while conflicting ordinary fields create separate rows.
    for row_result in rows:
        case_ref = row_result.get("Case reference")
        if not case_ref:
            continue
        aggregate_columns = {
            rule["template_column"]
            for rule in mapping
            if rule["append"] or rule["aggregate_by_case"]
        }
        non_append = {
            col: val for col, val in row_result.items()
            if col not in aggregate_columns and col != "Case reference"
        }
        if case_ref not in case_rows:
            case_rows[case_ref] = [{"values": {"Case reference": case_ref}, "conflicts": set(), "arrays": {}}]

        # Reuse a row when all non-array values agree, otherwise retain the
        # source combination in a new row and highlight the differing cells.
        compatible = None
        for record in case_rows[case_ref]:
            if all(
                col not in record["values"] or record["values"][col] == val
                for col, val in non_append.items()
            ):
                compatible = record
                break
        if compatible is None:
            compatible = {"values": {"Case reference": case_ref}, "conflicts": set(), "arrays": {}}
            for col, val in non_append.items():
                for record in case_rows[case_ref]:
                    if col in record["values"] and record["values"][col] != val:
                        record["conflicts"].add(col)
                compatible["values"][col] = val
                compatible["conflicts"].add(col)
            case_rows[case_ref].append(compatible)
        else:
            for col, val in non_append.items():
                compatible["values"].setdefault(col, val)

        for col, val in row_result.items():
            if col not in aggregate_columns or is_blank(val):
                continue
            compatible["arrays"].setdefault(col, [])
            if val not in compatible["arrays"][col]:
                compatible["arrays"][col].append(val)

    print(f"  {filename}: {len(mapping)} mapped columns, {len(unmapped_cols)} unmapped, {len(df)} rows")

all_rows = []
all_conflicts = []
for records in case_rows.values():
    for record in records:
        row = dict(record["values"])
        for col, values in record["arrays"].items():
            if len(values) == 1:
                row[col] = values[0]
            elif values:
                if col in _extended_data_columns:
                    row[col] = "[" + ", ".join(values) + "]"
                else:
                    row[col] = "{" + ", ".join(f"{index}: {value}" for index, value in enumerate(values, 1)) + "}"
        all_rows.append(row)
        all_conflicts.append(record["conflicts"])
print(f"\nBlock 4 done - {len(all_rows)} unique cases, {len(all_audit)} audit flags, {len(unmapped_report)} unmapped column entries")


# SPOT CHECK - show the first case
if all_rows:
    print("Block 5 spot check - first case:")
    for k, v in all_rows[0].items():
        print(f"  {k!r}: {v!r}")


# WRITE OUTPUT
os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
shutil.copy(MASTER_FILE, OUTPUT_FILE)   # never modify the original

wb = openpyxl.load_workbook(OUTPUT_FILE)
ws = wb[TEMPLATE_SHEET]

# Build column-name -> Excel column number from the header row
column_lookup = {}
for col_num in range(1, ws.max_column + 1):
    header_value = ws.cell(row=TEMPLATE_HEADER_ROW, column=col_num).value
    if header_value:
        column_lookup[str(header_value).strip()] = col_num

excel_row = TEMPLATE_FIRST_DATA_ROW
for row_result, conflicts in zip(all_rows, all_conflicts):
    for template_column, value in row_result.items():
        if template_column not in column_lookup:
            print(f"WARNING: '{template_column}' not found in Template headers - skipped")
            continue
        col_num = column_lookup[template_column]
        cell = ws.cell(row=excel_row, column=col_num)
        cell.value = value
        if isinstance(value, date):
            cell.number_format = "YYYY-MM-DD"
        if template_column in conflicts:
            cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FFF2CC")
    excel_row += 1

wb.save(OUTPUT_FILE)
print(f"Block 6 done - wrote {len(all_rows)} rows to {OUTPUT_FILE}")


# WRITE AUDIT LOG AND UNMAPPED COLUMNS REPORT
os.makedirs(os.path.dirname(AUDIT_LOG_FILE), exist_ok=True)

with open(AUDIT_LOG_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Case reference", "Template column", "Issue"])
    writer.writerows(all_audit)
print(f"Audit log:              {AUDIT_LOG_FILE} ({len(all_audit)} entries)")

with open(UNMAPPED_COLS_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["File", "Column"])
    writer.writeheader()
    writer.writerows(unmapped_report)
print(f"Unmapped columns report: {UNMAPPED_COLS_FILE} ({len(unmapped_report)} entries)")
