# %% [markdown]
# # Extract Horizon field names from the S62A discovery workbook
#
# Reads the "BO Crown Comparison - USE THIS" tab of S62A-Data-Discovery.xlsx and,
# for every row, works out:
#   - Field              -> column E  (the template field name)
#   - Source              -> column O  (which system(s) have this field)
#   - Horizon field name   -> derived from column P (Source field)
#   - Spreadsheet field name -> derived from column P (Source field)
#
# Example (row 3): Source = "Horizon/Spreadsheet", Source field =
# "Application description / Development description"
#   -> Horizon field name    = "Application description"
#   -> Spreadsheet field name = "Development description"
#
# Run block by block in VS Code's Interactive Window (Shift+Enter on each `# %%`).

# %%
# ============================================================
# BLOCK 1: CONFIG
# ============================================================
import re
import openpyxl
import csv
import os

DISCOVERY_FILE = "S62A-Data-Discovery.xlsx"
DISCOVERY_SHEET = "BO Crown Comparison - USE THIS"
OUTPUT_CSV = "output/horizon_field_mapping.csv"

print("Block 1 done - config set")

# %%
# ============================================================
# BLOCK 2: PARSING HELPERS
# ============================================================
def normalize_header(text):
    """Collapses whitespace/case so headers can be compared reliably."""
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().lower()

def find_column(header_row, predicate, description):
    for idx, header in enumerate(header_row, start=1):
        if predicate(normalize_header(header)):
            return idx
    raise ValueError(f"Discovery sheet: could not find a column for {description}")

def parse_horizon_and_spreadsheet_names(source_raw, source_field_raw):
    """Given the 'Source' and 'Source field' cells for one row, work out the
    Horizon-side and Spreadsheet-side field name. Either (or both) can be blank.
    When Source mentions both systems and the text has a '/', the order is
    always Horizon first, then Spreadsheet (confirmed with PINS): 'X / Y' ->
    Horizon = X, Spreadsheet = Y. If there's no '/', the same single name is
    used by both systems."""
    source_norm = normalize_header(source_raw)
    has_horizon = "horizon" in source_norm
    has_spreadsheet = "spreadsheet" in source_norm
    text = str(source_field_raw).strip() if source_field_raw not in (None, "") else ""

    horizon_name = spreadsheet_name = None
    if has_horizon and has_spreadsheet:
        if "/" in text:
            left, right = text.split("/", 1)
            horizon_name, spreadsheet_name = left.strip(), right.strip()
        else:
            horizon_name = spreadsheet_name = text
    elif has_horizon:
        horizon_name = text
    elif has_spreadsheet:
        spreadsheet_name = text
    return horizon_name, spreadsheet_name

# Free text that means "this isn't a clean single column name" (explanatory
# notes, TBC placeholders, etc.) - flagged rather than hidden, so you can
# manually review those rows.
MESSY_MARKERS = ("n/a", "tbc", "can't", "cannot", "no data", "?", "look at", "pull out")

def looks_clean(text):
    if not text:
        return False
    if len(text) > 80:
        return False
    return not any(marker in text.lower() for marker in MESSY_MARKERS)

print("Block 2 done - parsing helpers ready")

# %%
# ============================================================
# BLOCK 3: READ THE DISCOVERY SHEET AND BUILD THE MAPPING TABLE
# ============================================================
wb = openpyxl.load_workbook(DISCOVERY_FILE, data_only=True)
ws = wb[DISCOVERY_SHEET]
header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

field_col = find_column(header_row, lambda h: h == "field", "'Field'")
source_col = find_column(header_row, lambda h: h == "source", "'Source'")
source_field_col = find_column(header_row, lambda h: h.startswith("source field"),
                                "'Source field (Horizon / Spreadsheet)'")
confirmed_source_col = find_column(header_row, lambda h: h == "confirmed source", "'Confirmed source'")
confirmed_field_col = find_column(header_row, lambda h: h == "confirmed source field", "'Confirmed source field'")

mapping_rows = []
for row_num in range(2, ws.max_row + 1):
    field_name = ws.cell(row=row_num, column=field_col).value
    if field_name is None or not str(field_name).strip():
        continue

    # A manually "confirmed" source/field (if filled in) wins over the raw discovery notes.
    source_raw = ws.cell(row=row_num, column=confirmed_source_col).value or \
        ws.cell(row=row_num, column=source_col).value
    source_field_raw = ws.cell(row=row_num, column=confirmed_field_col).value or \
        ws.cell(row=row_num, column=source_field_col).value

    horizon_name, spreadsheet_name = parse_horizon_and_spreadsheet_names(source_raw, source_field_raw)

    mapping_rows.append({
        # template_field = column E - the column name in Template LEGACY cases S62A.xlsx
        # to write into. This is the key a downstream "map Horizon columns onto the
        # template" script should look up.
        "template_field": str(field_name).strip(),
        # horizon_column = the column name to look for in a Horizon export spreadsheet.
        # Blank means this template field has no known Horizon-side column.
        "horizon_column": horizon_name,
        # spreadsheet_column = the column name in the manual Section-62a-Cases spreadsheet
        # (kept for reference/cross-check, not needed by the Horizon mapping script).
        "spreadsheet_column": spreadsheet_name,
        # has_horizon_mapping = False if horizon_column is blank, OR if it looks like an
        # explanatory note rather than a real column name - review these by hand.
        "has_horizon_mapping": bool(horizon_name) and looks_clean(horizon_name),
        "source_type": source_raw,
        "discovery_row": row_num,
        "notes": source_field_raw,
    })

print(f"Block 3 done - parsed {len(mapping_rows)} rows from '{DISCOVERY_SHEET}'")

# %%
# ============================================================
# BLOCK 4: WRITE OUT A CSV YOU CAN REVIEW / HAND OFF
# ============================================================
os.makedirs(os.path.dirname(OUTPUT_CSV) or ".", exist_ok=True)
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=[
        "template_field", "horizon_column", "spreadsheet_column", "has_horizon_mapping",
        "source_type", "discovery_row", "notes",
    ])
    writer.writeheader()
    writer.writerows(mapping_rows)

with_horizon = sum(1 for r in mapping_rows if r["horizon_column"])
usable = sum(1 for r in mapping_rows if r["has_horizon_mapping"])
print(f"Block 4 done - wrote {len(mapping_rows)} rows to {OUTPUT_CSV}")
print(f"  {with_horizon} rows have a Horizon column name, {usable} of those look clean enough to "
      f"use automatically (has_horizon_mapping = True). Review the rest by hand.")
